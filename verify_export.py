#!/usr/bin/env python3
import openpyxl
from datetime import datetime

file_path = r"C:\Users\LUCAS CACAU\Downloads\Rotas_Otimizadas_2026-07-06.xlsx"

# Try to find and analyze any recent xlsx file
import os
import glob

downloads = glob.glob(os.path.expanduser("~/Downloads/*.xlsx"))
rotas_files = [f for f in downloads if "Rotas" in os.path.basename(f)]

if rotas_files:
    latest_file = max(rotas_files, key=os.path.getctime)
    print(f"📄 Analisando arquivo: {os.path.basename(latest_file)}")
    print(f"   Modificado: {datetime.fromtimestamp(os.path.getctime(latest_file))}\n")
    
    try:
        wb = openpyxl.load_workbook(latest_file)
        ws = wb['Clientes']
        
        print(f"📊 PLANILHA: {ws.title}")
        print(f"{'='*110}\n")
        
        # Print header
        header = []
        for col in range(1, 12):
            cell_value = ws.cell(row=1, column=col).value
            header.append(str(cell_value))
        print("HEADER:", " | ".join(header))
        print("-"*110)
        
        # Print rows
        for row_num in range(2, min(15, ws.max_row + 1)):
            row_data = []
            for col in range(1, 12):
                cell = ws.cell(row=row_num, column=col)
                value = cell.value
                row_data.append(value)
            
            if row_data[0]:  # Only if ID exists
                print(f"{str(row_data[0]):<10} | {str(row_data[1])[:30]:<30} | {str(row_data[2]):<10} | {str(row_data[3]):<6} | {str(row_data[4]):<10} | {str(row_data[5]):<4} | {str(row_data[6]):<4} | {str(row_data[7]):<4} | {str(row_data[8]):<4} | {str(row_data[9]):<4} | {str(row_data[10]):<4}")
        
        print(f"\n✅ Total de clientes na planilha: {ws.max_row - 1}")
        
    except Exception as e:
        print(f"❌ Erro ao ler arquivo: {e}")
else:
    print("❌ Nenhum arquivo Rotas_Otimizadas encontrado em Downloads")
    print("\nArquivos .xlsx em Downloads:")
    for f in downloads[:5]:
        print(f"  - {os.path.basename(f)}")
