#!/usr/bin/env python3
import openpyxl
import os

file_path = r"C:\Users\LUCAS CACAU\Downloads\Rotas_Otimizadas_2026-07-06.xlsx"
print(f"📄 Analisando arquivo: {os.path.basename(file_path)}\n")

# Load workbook
wb = openpyxl.load_workbook(file_path)
ws_clientes = wb['Clientes']

print(f"📊 PLANILHA: {ws_clientes.title}")
print(f"{'='*100}\n")

# Print header
header = []
for col in range(1, 12):
    cell_value = ws_clientes.cell(row=1, column=col).value
    header.append(str(cell_value))
print("HEADER:", header)
print()

# Print rows
print(f"{'ID':<10} {'NOME':<35} {'ROTA':<10} {'FREQ':<6} {'DUR':<10} {'SEG':<5} {'TER':<5} {'QUA':<5} {'QUI':<5} {'SEX':<5} {'SAB':<5}")
print("-"*100)

for row_num in range(2, min(15, ws_clientes.max_row + 1)):
    row_data = []
    for col in range(1, 12):
        cell = ws_clientes.cell(row=row_num, column=col)
        value = cell.value
        row_data.append(value)
    
    # Format for display
    client_id = str(row_data[0]) if row_data[0] else ""
    client_name = str(row_data[1])[:30] if row_data[1] else ""
    rota = str(row_data[2]) if row_data[2] else ""
    freq = str(row_data[3]) if row_data[3] else ""
    dur = str(row_data[4]) if row_data[4] else ""
    days = [str(d) if d else "" for d in row_data[5:11]]
    
    print(f"{client_id:<10} {client_name:<35} {rota:<10} {freq:<6} {dur:<10} {days[0]:<5} {days[1]:<5} {days[2]:<5} {days[3]:<5} {days[4]:<5} {days[5]:<5}")

print(f"\n✅ Total de clientes na planilha: {ws_clientes.max_row - 1}")
